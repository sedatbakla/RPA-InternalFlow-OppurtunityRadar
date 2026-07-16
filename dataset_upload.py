"""Read and validate user-provided flow datasets without persisting them."""

from __future__ import annotations

import csv
import hashlib
import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile

import pandas as pd

from import_data import prepare_flows


MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UPLOAD_MEGABYTES = MAX_UPLOAD_BYTES // (1024 * 1024)
MAX_DATASET_ROWS = 100_000
MAX_XLSX_EXPANDED_BYTES = 50 * 1024 * 1024
SUPPORTED_EXTENSIONS = (".csv", ".xlsx")


@dataclass(frozen=True)
class UploadedFlowDataset:
    """Validated uploaded flow data and display-safe file metadata."""

    flows: pd.DataFrame
    file_name: str
    file_format: str
    row_count: int
    column_count: int
    fingerprint: str
    export_stem: str


def safe_file_name(file_name: str) -> str:
    """Remove any path component from an untrusted uploaded file name."""
    normalized = str(file_name).replace("\\", "/")
    return normalized.rsplit("/", maxsplit=1)[-1] or "uploaded_dataset"


def safe_export_stem(file_name: str) -> str:
    """Return a filesystem-neutral stem for generated download names."""
    stem = Path(safe_file_name(file_name)).stem
    ascii_stem = (
        unicodedata.normalize("NFKD", stem)
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    safe_stem = re.sub(r"[^A-Za-z0-9_-]+", "_", ascii_stem).strip("_-")
    return (safe_stem[:80] or "uploaded_flows").lower()


def dataset_fingerprint(file_name: str, content: bytes) -> str:
    """Return a stable identity for one uploaded file name and payload."""
    digest = hashlib.sha256()
    digest.update(safe_file_name(file_name).encode("utf-8"))
    digest.update(b"\0")
    digest.update(content)
    return digest.hexdigest()


def _rows_to_dataframe(rows: list[list[object]], source_name: str) -> pd.DataFrame:
    """Build a DataFrame while preserving duplicate headers for validation."""
    non_empty_rows = [
        row
        for row in rows
        if any(value is not None and str(value).strip() for value in row)
    ]
    if not non_empty_rows:
        raise ValueError(f"{source_name} is empty")

    header = non_empty_rows[0]
    data_rows = non_empty_rows[1:]
    if not data_rows:
        raise ValueError(f"{source_name} contains no data rows")
    if len(data_rows) > MAX_DATASET_ROWS:
        raise ValueError(
            f"{source_name} exceeds the {MAX_DATASET_ROWS:,}-row limit"
        )

    invalid_rows = [
        index + 2
        for index, row in enumerate(data_rows)
        if len(row) != len(header)
    ]
    if invalid_rows:
        raise ValueError(
            f"{source_name} has inconsistent column counts at rows: "
            f"{invalid_rows[:10]}"
        )

    return pd.DataFrame(data_rows, columns=header)


def _read_csv_bytes(content: bytes, source_name: str) -> pd.DataFrame:
    """Read UTF-8 CSV bytes with comma or semicolon delimiter detection."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ValueError(
            f"{source_name} must use UTF-8 or UTF-8 BOM encoding"
        ) from error

    if not text.strip():
        raise ValueError(f"{source_name} is empty")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        delimiter = dialect.delimiter
    except csv.Error:
        first_line = next((line for line in text.splitlines() if line.strip()), "")
        delimiter = ";" if first_line.count(";") > first_line.count(",") else ","

    rows: list[list[object]] = []
    try:
        reader = csv.reader(StringIO(text), delimiter=delimiter, strict=True)
        for row in reader:
            rows.append(row)
            if len(rows) > MAX_DATASET_ROWS + 1:
                raise ValueError(
                    f"{source_name} exceeds the {MAX_DATASET_ROWS:,}-row limit"
                )
    except csv.Error as error:
        raise ValueError(f"{source_name} could not be parsed as CSV") from error

    return _rows_to_dataframe(rows, source_name)


def _read_xlsx_bytes(content: bytes, source_name: str) -> pd.DataFrame:
    """Read the first Excel worksheet without evaluating workbook formulas."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as error:
        raise RuntimeError("XLSX support is not installed") from error

    try:
        with ZipFile(BytesIO(content)) as archive:
            expanded_size = sum(item.file_size for item in archive.infolist())
    except BadZipFile as error:
        raise ValueError(f"{source_name} could not be read as an XLSX file") from error
    if expanded_size > MAX_XLSX_EXPANDED_BYTES:
        raise ValueError(
            f"{source_name} exceeds the expanded XLSX safety limit"
        )

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise ValueError(f"{source_name} could not be read as an XLSX file") from error

    try:
        if not workbook.sheetnames:
            raise ValueError(f"{source_name} does not contain a worksheet")

        worksheet = workbook[workbook.sheetnames[0]]
        rows: list[list[object]] = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) > MAX_DATASET_ROWS + 1:
                raise ValueError(
                    f"{source_name} exceeds the {MAX_DATASET_ROWS:,}-row limit"
                )
    finally:
        workbook.close()

    return _rows_to_dataframe(rows, source_name)


def load_uploaded_dataset(file_name: str, content: bytes) -> UploadedFlowDataset:
    """Parse, validate, and enrich one uploaded flow catalog in memory."""
    display_name = safe_file_name(file_name)
    if not content:
        raise ValueError("Uploaded file is empty")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"Uploaded file exceeds the {MAX_UPLOAD_MEGABYTES} MB size limit"
        )

    extension = Path(display_name).suffix.casefold()
    if extension not in SUPPORTED_EXTENSIONS:
        allowed = ", ".join(
            value.removeprefix(".").upper() for value in SUPPORTED_EXTENSIONS
        )
        raise ValueError(f"Unsupported file format. Use {allowed}")

    source_name = f"Uploaded dataset '{display_name}'"
    if extension == ".csv":
        raw = _read_csv_bytes(content, source_name)
        file_format = "CSV"
    else:
        raw = _read_xlsx_bytes(content, source_name)
        file_format = "XLSX"

    source_shape = raw.shape
    flows = prepare_flows(raw, source_name=source_name)
    return UploadedFlowDataset(
        flows=flows,
        file_name=display_name,
        file_format=file_format,
        row_count=source_shape[0],
        column_count=source_shape[1],
        fingerprint=dataset_fingerprint(display_name, content),
        export_stem=safe_export_stem(display_name),
    )
